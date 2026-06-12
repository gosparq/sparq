# -----------------------------------------------------------------------------
# sparQ - TimeEntry Model
# Copyright (c) 2025-2026 sparQ Software LLC. Licensed under AGPL-3.0.
# -----------------------------------------------------------------------------

"""Time entry model for tracking member work hours.

This module provides time tracking functionality including manual entries,
timer-based tracking, approval workflows, and billing integration.

Classes:
    TimeEntryStatus: Enum of time entry workflow states.
    TimeEntry: Member time record with job/task associations.

Example:
    Creating a time entry::

        from modules.base.presence.models.time_entry import TimeEntry

        entry = TimeEntry.create(
            member_id=member.id,
            date=date.today(),
            hours=4.5,
            description="Client meeting and follow-up",
            is_billable=True,
            job_id=job.id
        )

    Using the timer::

        entry = TimeEntry.start_timer(member_id=member.id)
        # ... later ...
        entry.stop_timer()

    Approval workflow (entries auto-submit on creation)::

        entry.approve(approved_by_id=manager.id)
"""

from datetime import date, datetime, timedelta
from decimal import Decimal
from enum import Enum
from typing import Any, TypedDict

from system.api.serialization import SerializableMixin
from system.db.database import db
from system.db.decorators import ModelRegistry
from system.db.mixins import AuditMixin
from system.db.workspace import OrganizationMixin
from system.db.raise_on_lazy import LAZY
from system.utils.calendar_utils import get_week_start


class _UnsetType:
    """Sentinel for 'no change' — distinguishes from None (clear field)."""
    def __repr__(self) -> str:
        return "<UNSET>"

_UNSET: _UnsetType = _UnsetType()


class BatchEntryData(TypedDict, total=False):
    """Typed dict for batch time entry creation via :meth:`TimeEntry.create_batch`."""

    date: date
    hours: float
    description: str | None
    category: str | None
    job_id: int | None
    is_billable: bool
    timer_start: datetime | None
    timer_end: datetime | None


class TimeEntryStatus(Enum):
    """Time entry approval workflow states.

    Attributes:
        SUBMITTED: Submitted for manager approval.
        APPROVED: Approved by manager.
        REJECTED: Rejected, needs correction.
        INVOICED: Included on a customer invoice.
    """

    SUBMITTED = "Submitted"
    APPROVED = "Approved"
    REJECTED = "Rejected"
    INVOICED = "Invoiced"


@ModelRegistry.register
class TimeEntry(db.Model, OrganizationMixin, AuditMixin, SerializableMixin):
    """Member time tracking record.

    Time entries track hours worked by members, optionally associated
    with jobs, visits, or tasks. Supports both manual entry and timer
    functionality.

    Attributes:
        member_id: Foreign key to OrganizationUser.
        date: Date of the work.
        hours: Hours worked (decimal).
        description: Description of work performed.
        is_billable: Whether hours are billable to customer.
        job_id: Optional associated Job.
        visit_id: Optional associated ScheduledVisit.
        task_id: Optional associated ScheduleTask.
        category: General category for non-job time.
        labor_cost_rate: Snapshot of member cost rate.
        bill_rate: Snapshot of member billing rate.
        labor_cost: Calculated cost (hours * labor_cost_rate).
        billing_amount: Calculated billing (hours * bill_rate).
        status: TimeEntryStatus workflow state.
        timer_start: Timer start time if using timer.
        timer_end: Timer end time.

    Relationships:
        member: Member who logged the time.
        job: Optional associated job.
        invoice: Invoice this entry was billed on.
    """

    __tablename__ = "time_entry"

    id = db.Column(db.Integer, primary_key=True)

    # --- Core fields (always required) ---
    member_id = db.Column(
        db.Integer, db.ForeignKey("organization_user.id", ondelete="CASCADE"), nullable=False, index=True
    )
    date = db.Column(db.Date, nullable=False, default=date.today)
    hours = db.Column(db.Numeric(5, 2), nullable=False)  # Max 999.99 hours
    description = db.Column(db.Text)
    is_billable = db.Column(db.Boolean, default=False, nullable=False)

    # --- Optional associations (nullable - use what applies) ---
    # job_id links to SERVICE module (optional dependency — no FK constraint)
    job_id = db.Column(db.Integer, nullable=True, index=True)
    # visit_id links to SERVICE module (optional dependency — no FK constraint)
    visit_id = db.Column(db.Integer, nullable=True)
    # task_id links to SERVICE module (optional dependency — no FK constraint)
    task_id = db.Column(db.Integer, nullable=True)
    category = db.Column(db.String(100))  # General category (e.g., "Admin", "Training")

    # --- Rates and costing (snapshot at time of entry) ---
    labor_cost_rate = db.Column(db.Numeric(10, 2))  # Cost rate (from member)
    bill_rate = db.Column(db.Numeric(10, 2))  # Billing rate (from member)
    labor_cost = db.Column(db.Numeric(10, 2))  # hours × labor_cost_rate (calculated)
    billing_amount = db.Column(db.Numeric(10, 2))  # hours × bill_rate (if billable)

    # --- Workflow ---
    status = db.Column(db.Enum(TimeEntryStatus), default=TimeEntryStatus.SUBMITTED, nullable=False)
    submitted_at = db.Column(db.DateTime, nullable=True)
    approved_by_id = db.Column(
        db.Integer, db.ForeignKey("organization_user.id", ondelete="SET NULL"), nullable=True
    )
    approved_at = db.Column(db.DateTime, nullable=True)
    rejected_reason = db.Column(db.Text)
    # invoice_id links to BILLING module (optional dependency — no FK constraint)
    invoice_id = db.Column(db.Integer, nullable=True)

    # Timer functionality
    timer_start = db.Column(db.DateTime, nullable=True)  # For running timer
    timer_end = db.Column(db.DateTime, nullable=True)

    # --- Timestamps ---
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    # created_by_id, updated_by_id from AuditMixin

    # --- Relationships ---
    member = db.relationship("OrganizationUser", backref=db.backref("time_entries", lazy=LAZY), foreign_keys=[member_id], lazy=LAZY)
    approved_by = db.relationship("OrganizationUser", backref=db.backref("approved_time_entries", lazy=LAZY), foreign_keys=[approved_by_id], lazy=LAZY)

    @property
    def job(self):
        """Optional Job association (requires Service module)."""
        if not self.job_id:
            return None
        from system.module.registry import module_enabled
        if not module_enabled("Service"):
            return None
        from modules.base.service.models.job import Job
        return Job.query.get(self.job_id)

    @property
    def invoice(self):
        """Optional Invoice association (requires Billing module)."""
        if not self.invoice_id:
            return None
        from system.module.registry import module_enabled
        if not module_enabled("Billing"):
            return None
        from modules.base.billing.models.invoice import Invoice
        return Invoice.query.get(self.invoice_id)

    # --- Indexes ---
    __table_args__ = (
        db.Index("ix_time_entry_member_date", "member_id", "date"),
        db.Index("ix_time_entry_status", "status"),
    )

    # --- Properties ---
    @property
    def is_timer_running(self) -> bool:
        """Check if timer is currently running"""
        return self.timer_start is not None and self.timer_end is None

    @property
    def is_from_clock_punch(self) -> bool:
        """Check if this entry was auto-created from a clock punch.

        Returns:
            True if a ClockPunch record is linked to this entry.
        """
        return self.clock_punch is not None

    @property
    def all_punches_pending_correction(self) -> bool:
        """Check if all punches for this clock entry have pending corrections.

        Returns:
            True if both IN and OUT punches have pending correction requests,
            meaning the member has already requested edits for everything.
        """
        if not self.is_from_clock_punch:
            return False
        from .punch_correction_request import PunchCorrectionRequest

        out_punch = self.clock_punch
        if not out_punch:
            return False
        in_punch = out_punch.get_matching_in()
        out_pending = PunchCorrectionRequest.get_pending_for_punch(out_punch.id) is not None
        in_pending = (
            PunchCorrectionRequest.get_pending_for_punch(in_punch.id) is not None
            if in_punch
            else True  # No IN punch means nothing left to correct
        )
        return in_pending and out_pending

    @property
    def timer_duration(self) -> float:
        """Get current timer duration in hours"""
        if not self.timer_start:
            return 0

        end_time = self.timer_end or datetime.utcnow()
        duration_seconds = (end_time - self.timer_start).total_seconds()
        return round(duration_seconds / 3600, 2)  # Convert to hours

    # --- Class Methods (CRUD) ---
    @classmethod
    def create(
        cls,
        member_id: int,
        date: date,
        hours: float,
        description: str | None = None,
        is_billable: bool = False,
        job_id: int | None = None,
        visit_id: int | None = None,
        task_id: int | None = None,
        category: str | None = None,
        timer_start: datetime | None = None,
        timer_end: datetime | None = None,
    ) -> "TimeEntry":
        """Create a new time entry with auto-populated rates"""
        from modules.base.core.models.organization_user import OrganizationUser

        member = OrganizationUser.get_by_id(member_id)
        if not member:
            raise ValueError(f"Member {member_id} not found")

        # Validate hours
        cls.validate_hours_range(hours)
        if category != "Clock":
            cls.validate_daily_total(member_id, date, hours)
        if timer_start and timer_end:
            cls.validate_time_overlap(member_id, date, timer_start, timer_end)

        # Auto-populate rates from member
        labor_cost_rate = member.labor_cost_rate or Decimal("0.00")
        bill_rate = member.bill_rate or Decimal("0.00")

        # Calculate costs
        labor_cost = Decimal(str(hours)) * labor_cost_rate
        billing_amount = Decimal(str(hours)) * bill_rate if is_billable else Decimal("0.00")

        entry = cls(
            member_id=member_id,
            date=date,
            hours=hours,
            description=description,
            is_billable=is_billable,
            job_id=job_id,
            visit_id=visit_id,
            task_id=task_id,
            category=category,
            labor_cost_rate=labor_cost_rate,
            bill_rate=bill_rate,
            labor_cost=labor_cost,
            billing_amount=billing_amount,
            timer_start=timer_start,
            timer_end=timer_end,
            submitted_at=datetime.utcnow(),
        )

        db.session.add(entry)
        db.session.commit()
        return entry

    @classmethod
    def create_batch(
        cls,
        member_id: int,
        entries: list[BatchEntryData],
    ) -> list["TimeEntry"]:
        """Create multiple time entries in a single transaction.

        Args:
            member_id: The member to create entries for.
            entries: List of dicts, each with keys:
                date (date), hours (float), description (str|None),
                category (str|None), job_id (int|None),
                is_billable (bool), timer_start (datetime|None),
                timer_end (datetime|None).

        Returns:
            List of created TimeEntry instances.

        Raises:
            ValueError: If member not found or validation fails.
        """
        from modules.base.core.models.organization_user import OrganizationUser

        member = OrganizationUser.get_by_id(member_id)
        if not member:
            raise ValueError(f"Member {member_id} not found")

        # Fetch rates once for the batch
        labor_cost_rate = member.labor_cost_rate or Decimal("0.00")
        bill_rate = member.bill_rate or Decimal("0.00")

        # Track hours and time ranges accumulated within this batch per date
        batch_hours_by_date: dict[date, float] = {}
        batch_ranges_by_date: dict[date, list[tuple[datetime, datetime]]] = {}

        created = []
        for entry_data in entries:
            hours = entry_data["hours"]
            entry_date = entry_data["date"]

            # Validate single-entry range
            cls.validate_hours_range(hours)

            # Validate daily total (include batch-accumulated hours)
            batch_accumulated = batch_hours_by_date.get(entry_date, 0)
            cls.validate_daily_total(
                member_id, entry_date, hours, batch_hours=batch_accumulated
            )

            # Validate time overlap
            entry_timer_start = entry_data.get("timer_start")
            entry_timer_end = entry_data.get("timer_end")
            if entry_timer_start and entry_timer_end:
                cls.validate_time_overlap(
                    member_id, entry_date, entry_timer_start, entry_timer_end
                )
                # Check against other entries in this batch
                for existing_start, existing_end in batch_ranges_by_date.get(entry_date, []):
                    if entry_timer_start < existing_end and entry_timer_end > existing_start:
                        raise ValueError(
                            "Time range overlaps with another entry in this submission"
                        )
                batch_ranges_by_date.setdefault(entry_date, []).append(
                    (entry_timer_start, entry_timer_end)
                )

            # Track batch hours for this date
            batch_hours_by_date[entry_date] = batch_accumulated + hours

            is_billable = entry_data.get("is_billable", False)
            labor_cost = Decimal(str(hours)) * labor_cost_rate
            billing_amount = Decimal(str(hours)) * bill_rate if is_billable else Decimal("0.00")

            entry = cls(
                member_id=member_id,
                date=entry_date,
                hours=hours,
                description=entry_data.get("description"),
                is_billable=is_billable,
                job_id=entry_data.get("job_id"),
                category=entry_data.get("category"),
                labor_cost_rate=labor_cost_rate,
                bill_rate=bill_rate,
                labor_cost=labor_cost,
                billing_amount=billing_amount,
                timer_start=entry_timer_start,
                timer_end=entry_timer_end,
                submitted_at=datetime.utcnow(),
            )
            db.session.add(entry)
            created.append(entry)

        db.session.commit()
        return created

    def delete(self) -> None:
        """Delete this time entry.

        Raises:
            ValueError: If the entry is approved or invoiced.
        """
        if self.status in (TimeEntryStatus.APPROVED, TimeEntryStatus.INVOICED):
            raise ValueError("Approved or invoiced entries cannot be deleted")
        db.session.delete(self)
        db.session.commit()

    def update_fields(
        self,
        hours: float,
        category: str | None,
        job_id: int | None,
        description: str | None,
        is_billable: bool,
        timer_start: datetime | None | _UnsetType = _UNSET,
        timer_end: datetime | None | _UnsetType = _UNSET,
        updated_by_id: int | None = None,
    ) -> None:
        """Update editable fields on a manual time entry.

        Args:
            hours: New hours value.
            category: New category string.
            job_id: New job association (or None to clear).
            description: New description text.
            is_billable: New billable flag.
            timer_start: New timer start (None clears, _UNSET = no change).
            timer_end: New timer end (None clears, _UNSET = no change).
            updated_by_id: ID of the user performing the edit (AuditMixin).

        Returns:
            None

        Raises:
            ValueError: If entry is approved or invoiced.
            ValueError: If entry is from a clock punch.
        """
        if self.status in (TimeEntryStatus.APPROVED, TimeEntryStatus.INVOICED):
            raise ValueError("Approved or invoiced entries cannot be edited")
        if self.is_from_clock_punch:
            raise ValueError("Clock punch entries must be edited via punch correction")

        # Validate hours
        self.validate_hours_range(hours)
        self.validate_daily_total(
            self.member_id, self.date, hours, exclude_entry_id=self.id
        )
        effective_start = timer_start if not isinstance(timer_start, _UnsetType) else self.timer_start
        effective_end = timer_end if not isinstance(timer_end, _UnsetType) else self.timer_end
        if effective_start and effective_end:
            self.validate_time_overlap(
                self.member_id, self.date,
                effective_start, effective_end,
                exclude_entry_id=self.id,
            )

        if updated_by_id is not None:
            self.updated_by_id = updated_by_id

        self.hours = Decimal(str(hours))
        self.category = category
        self.job_id = job_id
        self.description = description
        self.is_billable = is_billable

        if timer_start is not _UNSET:
            self.timer_start = timer_start
        if timer_end is not _UNSET:
            self.timer_end = timer_end

        self.recalculate_costs()

        # Re-submit rejected entries
        if self.status == TimeEntryStatus.REJECTED:
            self.status = TimeEntryStatus.SUBMITTED
            self.rejected_reason = None
            self.submitted_at = datetime.utcnow()

        db.session.commit()

    @classmethod
    def create_from_visit(cls, visit_id, member_id):
        """Create time entry from completed ScheduledVisit (requires SERVICE module)"""
        from system.module.registry import module_enabled

        if not module_enabled("Service"):
            raise ValueError("Service module is not enabled")

        from modules.base.service.models.scheduled_visit import ScheduledVisit

        visit = ScheduledVisit.scoped().get(visit_id)
        if not visit:
            raise ValueError(f"Visit {visit_id} not found")

        if not visit.actual_start or not visit.actual_end:
            raise ValueError("Visit must have actual start and end times")

        # Calculate hours from visit times
        duration = visit.actual_end - visit.actual_start
        hours = duration.total_seconds() / 3600

        return cls.create(
            member_id=member_id,
            date=visit.actual_start.date(),
            hours=round(hours, 2),
            description=f"Service visit: {visit.job.title}" if visit.job else "Service visit",
            is_billable=True,  # Default for client work
            job_id=visit.job_id,
            visit_id=visit_id,
        )

    @classmethod
    def get_for_date(cls, member_id: int, entry_date: date) -> list["TimeEntry"]:
        """Get time entries for a member on a specific date, ordered by creation."""
        from sqlalchemy.orm import joinedload

        return cls.scoped().options(
            joinedload(cls.clock_punch),
        ).filter_by(
            member_id=member_id, date=entry_date
        ).order_by(cls.created_at).all()

    @classmethod
    def get_total_hours_for_date(cls, member_id: int, entry_date: date) -> float:
        """Get total hours for a member on a specific date.

        Args:
            member_id: The member's ID.
            entry_date: The date to total.

        Returns:
            Total hours as a float, rounded to 2 decimal places.
        """
        entries = cls.get_for_date(member_id, entry_date)
        return round(sum(float(e.hours) for e in entries), 2)

    # --- Validation ---

    @staticmethod
    def validate_hours_range(hours: float) -> None:
        """Validate that hours are within the allowed range (0, 24].

        Args:
            hours: The hours value to validate.

        Raises:
            ValueError: If hours are not between 0 (exclusive) and 24 (inclusive).
        """
        if hours <= 0 or hours > 24:
            raise ValueError("Hours must be between 0 and 24")

    @classmethod
    def validate_daily_total(
        cls,
        member_id: int,
        entry_date: date,
        new_hours: float,
        exclude_entry_id: int | None = None,
        batch_hours: float = 0,
    ) -> None:
        """Validate that adding new_hours won't exceed 24 hours for the day.

        Args:
            member_id: The member's ID.
            entry_date: The date to check.
            new_hours: Hours being added or updated to.
            exclude_entry_id: Entry ID to exclude (for updates).
            batch_hours: Hours already accumulated in the current batch
                for this member+date (not yet committed to DB).

        Raises:
            ValueError: If total daily hours would exceed 24.
        """
        from sqlalchemy import func

        query = db.session.query(func.coalesce(func.sum(cls.hours), 0)).filter(
            cls.member_id == member_id,
            cls.date == entry_date,
        )
        if exclude_entry_id is not None:
            query = query.filter(cls.id != exclude_entry_id)

        existing_hours = float(query.scalar())
        total = existing_hours + float(batch_hours) + float(new_hours)

        if total > 24:
            raise ValueError("Daily hours cannot exceed 24")

    @classmethod
    def validate_time_overlap(
        cls,
        member_id: int,
        entry_date: date,
        timer_start: datetime,
        timer_end: datetime,
        exclude_entry_id: int | None = None,
    ) -> None:
        """Check that a time range does not overlap with existing entries.

        Args:
            member_id: The member's ID.
            entry_date: The date of the entry.
            timer_start: Start of the proposed time range (UTC).
            timer_end: End of the proposed time range (UTC).
            exclude_entry_id: Entry ID to exclude (for updates).

        Raises:
            ValueError: If the time range overlaps with an existing entry.
        """
        query = cls.scoped().filter(
            cls.member_id == member_id,
            cls.date == entry_date,
            cls.timer_start.isnot(None),
            cls.timer_end.isnot(None),
            cls.timer_start < timer_end,
            cls.timer_end > timer_start,
        )
        if exclude_entry_id is not None:
            query = query.filter(cls.id != exclude_entry_id)

        if query.first():
            raise ValueError(
                "Time range overlaps with an existing entry on this day"
            )

    @classmethod
    def get_by_member(cls, member_id, start_date=None, end_date=None):
        """Get time entries for a member, optionally filtered by date range"""
        query = cls.scoped().filter_by(member_id=member_id)

        if start_date:
            query = query.filter(cls.date >= start_date)
        if end_date:
            query = query.filter(cls.date <= end_date)

        return query.order_by(cls.date.desc()).all()

    @classmethod
    def get_for_user(cls, member_id: int):
        """Return a base query for a user's time entries, ordered by date desc.

        Returns an unpaginated query suitable for use with paginated_response().
        """
        return cls.scoped().filter_by(member_id=member_id).order_by(cls.date.desc())

    @classmethod
    def get_pending_approval(cls, through_date: date | None = None) -> list["TimeEntry"]:
        """Get all time entries awaiting approval.

        Excludes entries with a running timer (timer_start set, timer_end null)
        since they show 0 hours and shouldn't appear on the approval page.

        Args:
            through_date: If provided, only return entries on or before this date.

        Returns:
            List of submitted time entries, ordered by submission date descending.
        """
        from sqlalchemy.orm import joinedload
        from modules.base.core.models.organization_user import OrganizationUser
        query = cls.scoped().options(
            joinedload(cls.member).joinedload(OrganizationUser.user),
            joinedload(cls.clock_punch),
        ).filter_by(status=TimeEntryStatus.SUBMITTED).filter(
            db.or_(cls.timer_start.is_(None), cls.timer_end.isnot(None))
        )
        if through_date:
            query = query.filter(cls.date <= through_date)
        return query.order_by(cls.submitted_at.desc()).all()

    @classmethod
    def submitted_count(cls) -> int:
        """Count submitted time entries awaiting approval."""
        return cls.scoped().filter_by(status=TimeEntryStatus.SUBMITTED).count()

    @classmethod
    def get_entries_grouped_by_job(
        cls, member_id: int, week_start: date, week_end: date
    ) -> dict[str, dict[str, Any]]:
        """Get time entries for a week grouped by job/category.

        Args:
            member_id: ID of the member.
            week_start: Start date of the week.
            week_end: End date of the week.

        Returns:
            dict keyed by job/category with label, entries by date, and total hours.
        """
        entries = cls.scoped().filter(
            cls.member_id == member_id,
            cls.date >= week_start,
            cls.date <= week_end,
        ).all()

        entries_by_job: dict[str, dict[str, Any]] = {}
        for entry in entries:
            if entry.job_id and entry.job:
                key = f"job_{entry.job_id}"
                label = f"Job #{entry.job.job_number} - {entry.job.title}"
            elif entry.category and entry.category != "General":
                key = f"category_{entry.category}"
                label = entry.category
            else:
                key = "general"
                label = "General"

            if key not in entries_by_job:
                entries_by_job[key] = {"label": label, "entries": {}, "total": 0}

            date_key = entry.date.isoformat()
            if date_key not in entries_by_job[key]["entries"]:
                entries_by_job[key]["entries"][date_key] = []
            entries_by_job[key]["entries"][date_key].append(entry)
            entries_by_job[key]["total"] += float(entry.hours)

        return entries_by_job

    @classmethod
    def get_pending_grouped_by_member(cls, through_date: date | None = None) -> dict[int, dict[str, Any]]:
        """Get submitted time entries grouped by member for approval.

        Args:
            through_date: If provided, only include entries on or before this date.

        Returns:
            dict keyed by member_id with member, entries list, and total_hours.
        """
        pending_entries = cls.get_pending_approval(through_date=through_date)

        entries_by_member: dict[int, dict[str, Any]] = {}
        for entry in pending_entries:
            if entry.member_id not in entries_by_member:
                entries_by_member[entry.member_id] = {
                    "member": entry.member,
                    "entries": [],
                    "total_hours": 0,
                }
            entries_by_member[entry.member_id]["entries"].append(entry)
            entries_by_member[entry.member_id]["total_hours"] += float(entry.hours)

        return entries_by_member

    @classmethod
    def get_approved_unbilled(cls):
        """Get approved billable time entries not yet invoiced"""
        return (
            cls.scoped().filter_by(
                status=TimeEntryStatus.APPROVED, is_billable=True, invoice_id=None
            )
            .order_by(cls.date.desc())
            .all()
        )

    @classmethod
    def get_payroll_summary(
        cls,
        week_start: date,
        week_end: date,
        member_id: int | None = None,
        expense_totals: dict[int, float] | None = None,
    ) -> dict:
        """Get approved time entries grouped by member for payroll.

        Args:
            week_start: Start date of the pay period.
            week_end: End date of the pay period.
            member_id: Optional member ID to filter to single member.
            expense_totals: Optional dict mapping user_id to reimbursable expense
                total. When provided, each member dict is enriched with
                ``total_expenses``.

        Returns:
            dict: Dictionary keyed by member_id with member, total_hours, total_cost,
                and optionally total_expenses.
        """
        query = cls.scoped().filter(
            cls.status == TimeEntryStatus.APPROVED,
            cls.date >= week_start,
            cls.date <= week_end,
        )

        if member_id:
            query = query.filter(cls.member_id == member_id)

        entries = query.all()

        # Group by member
        summary = {}
        for entry in entries:
            if entry.member_id not in summary:
                summary[entry.member_id] = {
                    "member": entry.member,
                    "total_hours": 0,
                    "total_cost": 0,
                }
            summary[entry.member_id]["total_hours"] += float(entry.hours)
            summary[entry.member_id]["total_cost"] += float(entry.labor_cost or 0)

        if expense_totals is not None:
            for data in summary.values():
                data["total_expenses"] = expense_totals.get(data["member"].user_id, 0)

        return summary

    @classmethod
    def export_payroll_csv(cls, week_start: date, week_end: date, member_id: int | None = None) -> tuple[str, str]:
        """Export payroll data as CSV string with filename.

        Args:
            week_start: Start date of the pay period.
            week_end: End date of the pay period.
            member_id: Optional member ID to filter to single member.

        Returns:
            tuple: (csv_content: str, filename: str)
        """
        import csv
        from io import StringIO

        summary = cls.get_payroll_summary(week_start, week_end, member_id)

        output = StringIO()
        writer = csv.writer(output)

        # Header row
        writer.writerow(["Employee ID", "Employee Name", "Total Hours", "Total Cost", "Status"])

        # Data rows
        for data in summary.values():
            writer.writerow([
                data["member"].employee_id,
                f"{data['member'].user.first_name} {data['member'].user.last_name}",
                f"{data['total_hours']:.2f}",
                f"{data['total_cost']:.2f}",
                "Approved",
            ])

        # Summary row (only if multiple members)
        if len(summary) > 1:
            total_hours = sum(d["total_hours"] for d in summary.values())
            total_cost = sum(d["total_cost"] for d in summary.values())
            writer.writerow([])
            writer.writerow(["TOTAL", "", f"{total_hours:.2f}", f"{total_cost:.2f}", ""])

        # Generate filename
        if member_id and summary:
            emp = list(summary.values())[0]["member"]
            filename = f"{emp.user.first_name.lower()}_{emp.user.last_name.lower()}_{week_end}.csv"
        else:
            filename = f"payroll_{week_end}.csv"

        return output.getvalue(), filename

    @classmethod
    def get_trailing_two_week_range(cls) -> tuple[date, date, date, date]:
        """Get date ranges for trailing two weeks (one week behind current).

        The trailing two weeks are the two most recently completed weeks,
        not including the current week. This gives time for timesheet
        approvals before data appears in the export.

        Example: On Friday Jan 17th, returns:
            Week 1: Dec 30 - Jan 5
            Week 2: Jan 6 - Jan 12
            (Current week Jan 13-19 is excluded)

        Returns:
            tuple: (week1_start, week1_end, week2_start, week2_end)
        """
        today = date.today()
        # Get first day of current week (respects company setting)
        current_week_start = get_week_start(today)

        # Week 2 is last week (one week behind)
        week2_start = current_week_start - timedelta(days=7)
        week2_end = week2_start + timedelta(days=6)

        # Week 1 is two weeks behind
        week1_start = current_week_start - timedelta(days=14)
        week1_end = week1_start + timedelta(days=6)

        return week1_start, week1_end, week2_start, week2_end

    @classmethod
    def get_daily_breakdown(cls, week_start: date, week_end: date) -> dict:
        """Get approved hours per day per member for a week.

        Args:
            week_start: First day of the week.
            week_end: Last day of the week.

        Returns:
            dict: {member_id: {"member": WorkspaceUser, "days": {date: hours}, "total": float}}
        """
        entries = cls.scoped().filter(
            cls.status == TimeEntryStatus.APPROVED,
            cls.date >= week_start,
            cls.date <= week_end,
        ).all()

        # Initialize result structure
        breakdown = {}
        for entry in entries:
            emp_id = entry.member_id
            if emp_id not in breakdown:
                breakdown[emp_id] = {
                    "member": entry.member,
                    "days": {},
                    "total": 0.0,
                }

            # Add hours to day
            entry_date = entry.date
            if entry_date not in breakdown[emp_id]["days"]:
                breakdown[emp_id]["days"][entry_date] = 0.0
            breakdown[emp_id]["days"][entry_date] += float(entry.hours)
            breakdown[emp_id]["total"] += float(entry.hours)

        return breakdown

    @staticmethod
    def _build_week_sheet(
        ws: Any,
        start_row: int,
        week_start: date,
        week_end: date,
        data: dict[int, Any],
    ) -> int:
        """Write a week's daily-breakdown table into an Excel worksheet.

        Renders a table with member rows, day-of-week columns, and totals.
        Includes a grand-total row when multiple members are present.

        Args:
            ws: openpyxl Worksheet to write into.
            start_row: First row number for this table.
            week_start: First day of the week.
            week_end: Last day of the week.
            data: Output of ``get_daily_breakdown()`` — keyed by member_id.

        Returns:
            Next available row number (with spacing for subsequent tables).
        """
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from system.i18n.translation import translate as _

        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        row = start_row

        # Week header
        week_title = _("Week") + f": {week_start.strftime('%b %d')} - {week_end.strftime('%b %d, %Y')}"
        ws.cell(row=row, column=1, value=week_title).font = header_font
        row += 1

        # Generate day columns
        days = []
        current_day = week_start
        while current_day <= week_end:
            days.append(current_day)
            current_day += timedelta(days=1)

        # Column headers
        headers = [_("Employee")] + [d.strftime("%a %m/%d") for d in days] + [_("Total")]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Data rows (sorted by member name)
        sorted_members = sorted(data.values(), key=lambda x: x["member"].user.last_name)
        for emp_data in sorted_members:
            emp = emp_data["member"]
            emp_name = f"{emp.user.first_name} {emp.user.last_name}"

            cell = ws.cell(row=row, column=1, value=emp_name)
            cell.border = thin_border

            for col, day in enumerate(days, 2):
                hours = emp_data["days"].get(day, 0)
                cell = ws.cell(row=row, column=col, value=hours if hours > 0 else "")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

            cell = ws.cell(row=row, column=len(headers), value=emp_data["total"])
            cell.border = thin_border
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

            row += 1

        # Grand total row if multiple members
        if len(data) > 1:
            cell = ws.cell(row=row, column=1, value=_("TOTAL"))
            cell.font = Font(bold=True)
            cell.border = thin_border

            for col, day in enumerate(days, 2):
                day_total = sum(emp["days"].get(day, 0) for emp in data.values())
                cell = ws.cell(row=row, column=col, value=day_total if day_total > 0 else "")
                cell.border = thin_border
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            grand_total = sum(emp["total"] for emp in data.values())
            cell = ws.cell(row=row, column=len(headers), value=grand_total)
            cell.border = thin_border
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            row += 1

        return row + 2  # Add spacing before next table

    @classmethod
    def export_member_excel(
        cls,
        week_start: date,
        week_end: date,
        member_id: int,
        expense_items: list[Any] | None = None,
    ) -> tuple[bytes, str]:
        """Export a single member's daily breakdown for one week as Excel.

        Args:
            week_start: First day of the week to export.
            week_end: Last day of the week to export.
            member_id: ID of the member to export.
            expense_items: Optional list of Expense objects for itemized display.

        Returns:
            Tuple of (excel_bytes, filename).
        """
        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from system.i18n.translation import translate as _

        # Get daily breakdown filtered to this member
        all_data = cls.get_daily_breakdown(week_start, week_end)
        emp_data = {member_id: all_data[member_id]} if member_id in all_data else {}

        wb = Workbook()
        ws = wb.active
        ws.title = "Timesheet"

        # Title
        if emp_data:
            emp = emp_data[member_id]["member"]
            emp_name = f"{emp.user.first_name} {emp.user.last_name}"
        else:
            emp_name = _("Employee")

        title = _("Timesheet") + f": {emp_name} — {week_start.strftime('%b %d')} - {week_end.strftime('%b %d, %Y')}"
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:I1")

        cls._build_week_sheet(ws, 3, week_start, week_end, emp_data)

        # Add itemized reimbursable expenses table if provided
        if expense_items:
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Leave a gap after the hours table
            row = ws.max_row + 3

            # Section header
            ws.cell(row=row, column=1, value=_("Reimbursable Expenses")).font = Font(bold=True, size=12)
            row += 1

            # Column headers
            exp_headers = [_("Date"), _("Description"), _("Category"), _("Amount")]
            for col, header in enumerate(exp_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
            row += 1

            # Expense rows
            total_amount = 0
            for expense in expense_items:
                ws.cell(row=row, column=1, value=expense.expense_date.strftime("%m/%d/%Y")).border = thin_border
                ws.cell(row=row, column=2, value=expense.description).border = thin_border
                ws.cell(row=row, column=3, value=expense.category.value).border = thin_border
                amount_cell = ws.cell(row=row, column=4, value=f"${float(expense.amount):.2f}")
                amount_cell.border = thin_border
                amount_cell.alignment = Alignment(horizontal="right")
                total_amount += float(expense.amount)
                row += 1

            # Total row
            total_label = ws.cell(row=row, column=3, value=_("Total"))
            total_label.font = Font(bold=True)
            total_label.border = thin_border
            total_label.alignment = Alignment(horizontal="right")
            total_cell = ws.cell(row=row, column=4, value=f"${total_amount:.2f}")
            total_cell.font = Font(bold=True)
            total_cell.border = thin_border
            total_cell.alignment = Alignment(horizontal="right")

        # Column widths (A is shared between employee name and expense date)
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 12
        for col in range(5, 10):
            ws.column_dimensions[get_column_letter(col)].width = 12

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{emp_name.lower().replace(' ', '_')}_{week_end.strftime('%Y-%m-%d')}.xlsx"
        return output.getvalue(), filename

    @classmethod
    def export_payroll_excel(
        cls,
        start_date: date | None = None,
        end_date: date | None = None,
        expense_totals: dict[int, float] | None = None,
    ) -> tuple[bytes, str]:
        """Export payroll report as Excel with daily breakdown per week.

        Creates an Excel workbook with one table per week in the given range,
        showing daily breakdown of approved hours per employee.

        Args:
            start_date: First day of the export range. Defaults to trailing
                two-week range if not provided.
            end_date: Last day of the export range. Defaults to trailing
                two-week range if not provided.
            expense_totals: Optional dict of user_id -> reimbursable amount.

        Returns:
            Tuple of (excel_bytes, filename).
        """
        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
        from system.i18n.translation import translate as _

        # Default to trailing two weeks for backward compat
        if start_date is None or end_date is None:
            week1_start, _, _, week2_end = cls.get_trailing_two_week_range()
            start_date = week1_start
            end_date = week2_end

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Timesheets"

        title = _("Timesheets") + f": {start_date.strftime('%b %d, %Y')} - {end_date.strftime('%b %d, %Y')}"
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:I1")

        current_row = 3

        # Break range into weeks and write a table per week
        week_cursor = get_week_start(start_date)
        while week_cursor <= end_date:
            week_end = week_cursor + timedelta(days=6)
            # Clamp to the actual export range
            effective_start = max(week_cursor, start_date)
            effective_end = min(week_end, end_date)

            data = cls.get_daily_breakdown(effective_start, effective_end)
            current_row = cls._build_week_sheet(
                ws, current_row, effective_start, effective_end, data
            )

            week_cursor += timedelta(days=7)

        # Add reimbursable expenses summary if provided
        if expense_totals:
            from modules.base.core.models.user import User

            user_ids = list(expense_totals.keys())
            users = {
                u.id: f"{u.first_name} {u.last_name}"
                for u in User.get_by_ids(user_ids)
            }

            row = current_row
            ws.cell(row=row, column=1, value=_("Reimbursable Expenses")).font = Font(bold=True, size=12)
            row += 1
            ws.cell(row=row, column=1, value=_("Employee")).font = Font(bold=True)
            ws.cell(row=row, column=2, value=_("Amount")).font = Font(bold=True)
            row += 1

            grand_total = 0
            for uid, amount in sorted(expense_totals.items(), key=lambda x: users.get(x[0], "")):
                ws.cell(row=row, column=1, value=users.get(uid, f"User {uid}"))
                ws.cell(row=row, column=2, value=f"${amount:.2f}")
                grand_total += amount
                row += 1

            ws.cell(row=row, column=1, value=_("TOTAL")).font = Font(bold=True)
            ws.cell(row=row, column=2, value=f"${grand_total:.2f}").font = Font(bold=True)

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        for col in range(2, 10):
            ws.column_dimensions[get_column_letter(col)].width = 12

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"timesheet_{start_date.strftime('%m-%d')}_{end_date.strftime('%m-%d')}.xlsx"
        return output.getvalue(), filename

    @classmethod
    def export_payroll_csv_range(
        cls,
        start_date: date,
        end_date: date,
        member_id: int | None = None,
        expense_totals: dict[int, float] | None = None,
    ) -> tuple[str, str]:
        """Export payroll summary for a date range as CSV.

        Args:
            start_date: First day of the export range.
            end_date: Last day of the export range.
            member_id: Optional member ID to filter to a single member.
            expense_totals: Optional dict mapping user_id to reimbursable expense
                amount. When provided, an Expenses column is included.

        Returns:
            Tuple of (csv_content, filename).
        """
        import csv
        from io import StringIO
        from system.i18n.translation import translate as _

        summary = cls.get_payroll_summary(start_date, end_date, member_id)
        has_expenses = expense_totals is not None

        output = StringIO()
        writer = csv.writer(output)

        header = [_("Employee ID"), _("Employee Name")]
        if has_expenses:
            header.append(_("Expenses"))
        header.extend([_("Total Hours"), _("Total Cost"), _("Status")])
        writer.writerow(header)

        for data in summary.values():
            row = [
                data["member"].employee_id,
                f"{data['member'].user.first_name} {data['member'].user.last_name}",
            ]
            if has_expenses:
                emp_expenses = expense_totals.get(data["member"].user_id, 0)
                row.append(f"{emp_expenses:.2f}")
            row.extend([
                f"{data['total_hours']:.2f}",
                f"{data['total_cost']:.2f}",
                _("Approved"),
            ])
            writer.writerow(row)

        if len(summary) > 1:
            total_row = [_("TOTAL"), ""]
            if has_expenses:
                total_expenses = sum(expense_totals.get(d["member"].user_id, 0) for d in summary.values())
                total_row.append(f"{total_expenses:.2f}")
            total_hours = sum(d["total_hours"] for d in summary.values())
            total_cost = sum(d["total_cost"] for d in summary.values())
            total_row.extend([f"{total_hours:.2f}", f"{total_cost:.2f}", ""])
            writer.writerow([])
            writer.writerow(total_row)

        if member_id and summary:
            emp = list(summary.values())[0]["member"]
            filename = f"{emp.user.first_name.lower()}_{emp.user.last_name.lower()}_{end_date}.csv"
        else:
            filename = f"payroll_{start_date.strftime('%m-%d')}_{end_date.strftime('%m-%d')}.csv"
        return output.getvalue(), filename

    @classmethod
    def approve_all_for_member(
        cls, member_id: int, approved_by_id: int, through_date: date | None = None
    ) -> int:
        """Approve all submitted time entries for a member.

        Args:
            member_id: ID of the member whose entries to approve.
            approved_by_id: ID of the user performing the approval.
            through_date: If provided, only approve entries on or before this date.

        Returns:
            Number of entries approved.

        Raises:
            ValueError: If no pending entries found for member.
        """
        query = cls.scoped().filter_by(
            member_id=member_id,
            status=TimeEntryStatus.SUBMITTED
        )
        if through_date:
            query = query.filter(cls.date <= through_date)
        entries = query.all()

        if not entries:
            raise ValueError("No pending entries found for this member")

        now = datetime.utcnow()
        for entry in entries:
            entry.status = TimeEntryStatus.APPROVED
            entry.approved_by_id = approved_by_id
            entry.approved_at = now

        db.session.commit()
        return len(entries)

    # --- Instance Methods ---

    def approve(self, approved_by_id: int) -> None:
        """Approve time entry"""
        if self.status != TimeEntryStatus.SUBMITTED:
            raise ValueError("Only submitted entries can be approved")

        self.status = TimeEntryStatus.APPROVED
        self.approved_by_id = approved_by_id
        self.approved_at = datetime.utcnow()
        db.session.commit()

    def reject(self, reason: str, rejected_by_id: int) -> None:
        """Reject time entry with reason"""
        if self.status != TimeEntryStatus.SUBMITTED:
            raise ValueError("Only submitted entries can be rejected")

        self.status = TimeEntryStatus.REJECTED
        self.rejected_reason = reason
        db.session.commit()

    def mark_invoiced(self, invoice_id: int) -> None:
        """Mark entry as invoiced"""
        if self.status != TimeEntryStatus.APPROVED:
            raise ValueError("Only approved entries can be invoiced")
        if not self.is_billable:
            raise ValueError("Only billable entries can be invoiced")

        self.status = TimeEntryStatus.INVOICED
        self.invoice_id = invoice_id
        db.session.commit()

    def start_timer(self) -> None:
        """Start a timer for this entry"""
        if self.is_timer_running:
            raise ValueError("Timer is already running")

        self.timer_start = datetime.utcnow()
        self.timer_end = None
        db.session.commit()

    def stop_timer(self) -> None:
        """Stop the timer and update hours"""
        if not self.is_timer_running:
            raise ValueError("Timer is not running")

        self.timer_end = datetime.utcnow()
        self.hours = Decimal(str(self.timer_duration))

        # Recalculate costs
        self.labor_cost = self.hours * (self.labor_cost_rate or Decimal("0.00"))
        if self.is_billable:
            self.billing_amount = self.hours * (self.bill_rate or Decimal("0.00"))

        db.session.commit()

    def recalculate_costs(self) -> None:
        """Recalculate labor cost and billing amount"""
        self.labor_cost = Decimal(str(self.hours)) * (self.labor_cost_rate or Decimal("0.00"))
        if self.is_billable:
            self.billing_amount = Decimal(str(self.hours)) * (self.bill_rate or Decimal("0.00"))
        else:
            self.billing_amount = Decimal("0.00")
        db.session.commit()

    def recalculate_from_punches(self) -> None:
        """Recalculate hours from associated clock punches, applying rounding.

        When a clock punch is edited, this method recalculates the hours for
        the TimeEntry based on the clock-in and clock-out times. Applies time
        rounding if enabled in settings, matching the original clock_out calculation.

        Updates:
        - hours (with rounding if enabled)
        - description (with display times)
        - labor_cost
        - billing_amount (if billable)

        Note:
            This method expects to be called on a TimeEntry that has an
            associated clock_punch (the OUT punch). If not found, does nothing.
            Commits the database session.
        """
        import pytz
        from flask import g

        from .clock_punch import ClockPunch, PunchType
        from .settings import TimeTrackingSettings

        # Get the clock-out punch via the relationship
        clock_out_punch = self.clock_punch
        if not clock_out_punch:
            return

        # Find the matching clock-in punch (the most recent IN punch before this OUT)
        clock_in_punch = clock_out_punch.get_matching_in()
        if not clock_in_punch:
            return

        # Get times for calculation
        clock_in_time = clock_in_punch.punch_time
        clock_out_time = clock_out_punch.punch_time

        # Apply rounding if enabled (matching original clock_out logic)
        settings = TimeTrackingSettings.get()
        if settings.rounding_enabled:
            clock_in_time = ClockPunch.round_time(
                clock_in_time, PunchType.IN, settings.rounding_minutes, settings.rounding_type
            )
            clock_out_time = ClockPunch.round_time(
                clock_out_time, PunchType.OUT, settings.rounding_minutes, settings.rounding_type
            )

        # Calculate new hours
        duration = clock_out_time - clock_in_time
        hours = Decimal(str(duration.total_seconds() / 3600)).quantize(Decimal("0.01"))

        # Handle negative or zero hours
        if hours <= 0:
            hours = Decimal("0.25")  # Minimum 15 minutes

        # Update time entry
        self.hours = hours

        # Update description with LOCAL display times (not rounded times)
        company_settings = g.get("company_settings")
        tz_name = company_settings.timezone if company_settings else "America/Chicago"
        local_tz = pytz.timezone(tz_name)
        local_clock_in = clock_in_punch.punch_time.replace(tzinfo=pytz.UTC).astimezone(local_tz)
        local_clock_out = clock_out_punch.punch_time.replace(tzinfo=pytz.UTC).astimezone(local_tz)
        self.description = f"Clock: {local_clock_in.strftime('%I:%M %p')} - {local_clock_out.strftime('%I:%M %p')}"

        # Recalculate costs
        self.labor_cost = self.hours * (self.labor_cost_rate or Decimal("0.00"))
        if self.is_billable:
            self.billing_amount = self.hours * (self.bill_rate or Decimal("0.00"))

        db.session.commit()
